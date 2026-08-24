"""
Reads a "Timed Sequence" sheet into a validated list of pump steps.

Expected columns (matched loosely by header text, order doesn't matter):
    Time - Duration [h / min / s]  |  RPM  |  Motor State [ON/OFF]  |  Direction [CW/CCW]

Everything is validated BEFORE the pump is touched. A sequence that would
have thrown an error at step 27 of 32 fails here instead, at load time,
which is when you can still do something about it.
"""

from __future__ import annotations

from dataclasses import dataclass, field

from openpyxl import load_workbook

UNIT_SECONDS = {"s": 1, "sec": 1, "seconds": 1,
                "min": 60, "m": 60, "minutes": 60,
                "h": 3600, "hr": 3600, "hours": 3600}


@dataclass
class Step:
    index: int
    duration_s: float
    rpm: int
    motor_on: bool
    direction: str          # "CW" or "CCW"
    row: int                # spreadsheet row, for error messages

    @property
    def label(self) -> str:
        if not self.motor_on:
            return f"OFF for {fmt_duration(self.duration_s)}"
        return (f"{self.rpm} rpm {self.direction} "
                f"for {fmt_duration(self.duration_s)}")


@dataclass
class Sequence:
    steps: list[Step] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    source: str = ""
    sheet: str = ""
    unit: str = "min"

    @property
    def total_seconds(self) -> float:
        return sum(s.duration_s for s in self.steps)

    def elapsed_at_start_of(self, i: int) -> float:
        return sum(s.duration_s for s in self.steps[:i])


def fmt_duration(seconds: float) -> str:
    seconds = int(round(seconds))
    h, rem = divmod(seconds, 3600)
    m, s = divmod(rem, 60)
    if h:
        return f"{h}h {m:02d}m {s:02d}s"
    if m:
        return f"{m}m {s:02d}s"
    return f"{s}s"


def _norm(value) -> str:
    return "" if value is None else str(value).strip()


def _find_columns(rows) -> tuple[int, dict[str, int]]:
    """Locate the header row and map our four fields onto column indices."""
    wanted = {
        "duration": ("time", "duration"),
        "rpm": ("rpm", "speed"),
        "state": ("motor", "state", "on"),
        "direction": ("direction", "cw"),
    }
    for r_idx, row in enumerate(rows[:20]):
        cells = [_norm(c).lower() for c in row]
        found = {}
        for key, keywords in wanted.items():
            for c_idx, text in enumerate(cells):
                if text and any(k in text for k in keywords) and c_idx not in found.values():
                    found[key] = c_idx
                    break
        if len(found) == 4:
            return r_idx, found
    raise ValueError(
        "Could not find the header row. The sheet needs columns for duration, "
        "RPM, motor state and direction.")


def load_sequence(path: str, sheet: str | None = None, unit: str = "min",
                  min_rpm: int = 3, max_rpm: int = 400) -> Sequence:
    if unit not in UNIT_SECONDS:
        raise ValueError(f"Unknown time unit {unit!r}")
    multiplier = UNIT_SECONDS[unit]

    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet is None:
        # Prefer a sheet that looks like a sequence, else the first one.
        sheet = next((s for s in wb.sheetnames if "sequence" in s.lower()),
                     wb.sheetnames[0])
    ws = wb[sheet]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    header_row, cols = _find_columns(rows)
    seq = Sequence(source=path, sheet=sheet, unit=unit)

    blank_streak = 0
    for r_idx in range(header_row + 1, len(rows)):
        row = rows[r_idx]

        def cell(key):
            i = cols[key]
            return _norm(row[i]) if i < len(row) else ""

        dur_raw, rpm_raw = cell("duration"), cell("rpm")
        state_raw, dir_raw = cell("state"), cell("direction")
        excel_row = r_idx + 1

        # A completely empty row ends the sequence. Two in a row is definitive.
        # This is what stops us reading the stray dropdown-source cells that
        # are parked off to the right of the data.
        if not any((dur_raw, rpm_raw, state_raw, dir_raw)):
            blank_streak += 1
            if blank_streak >= 2:
                break
            continue
        blank_streak = 0

        try:
            duration = float(dur_raw)
        except ValueError:
            raise ValueError(f"Row {excel_row}: duration {dur_raw!r} is not a number")
        if duration <= 0:
            raise ValueError(f"Row {excel_row}: duration must be greater than zero")

        state = state_raw.upper()
        if state not in ("ON", "OFF"):
            raise ValueError(f"Row {excel_row}: motor state must be ON or OFF, got {state_raw!r}")
        motor_on = state == "ON"

        direction = dir_raw.upper().replace("ACW", "CCW")
        if direction not in ("CW", "CCW"):
            raise ValueError(f"Row {excel_row}: direction must be CW or CCW, got {dir_raw!r}")

        try:
            rpm = int(round(float(rpm_raw)))
        except ValueError:
            raise ValueError(f"Row {excel_row}: RPM {rpm_raw!r} is not a number")

        if motor_on and not (min_rpm <= rpm <= max_rpm):
            raise ValueError(
                f"Row {excel_row}: {rpm} rpm is outside the pump range "
                f"({min_rpm}-{max_rpm} rpm)")

        seq.steps.append(Step(
            index=len(seq.steps), duration_s=duration * multiplier,
            rpm=rpm, motor_on=motor_on, direction=direction, row=excel_row))

    if not seq.steps:
        raise ValueError("No usable steps found in the sheet")

    # Sanity warnings - not fatal, but you want to see them before you commit
    # 30-odd hours of fermentation to this.
    if seq.total_seconds > 72 * 3600:
        seq.warnings.append(
            f"Total run time is {fmt_duration(seq.total_seconds)}. "
            f"Check the time unit is really '{unit}'.")
    if seq.total_seconds < 300:
        seq.warnings.append(
            f"Total run time is only {fmt_duration(seq.total_seconds)}. "
            f"Check the time unit is really '{unit}'.")
    if len({s.direction for s in seq.steps}) > 1:
        seq.warnings.append("This sequence changes direction partway through.")
    return seq


if __name__ == "__main__":
    import sys
    s = load_sequence(sys.argv[1], unit=sys.argv[2] if len(sys.argv) > 2 else "min")
    print(f"{s.sheet}: {len(s.steps)} steps, total {fmt_duration(s.total_seconds)}")
    for w in s.warnings:
        print("  ! " + w)
    for st in s.steps:
        print(f"  {st.index + 1:>3}. row {st.row:<4} {st.label}")
